from pathlib import Path
from io import StringIO
from typing import Tuple
from matplotlib.widgets import RectangleSelector
from PyQt5 import QtWidgets
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from scipy import stats
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
from mpl_toolkits.basemap import Basemap
from datetime import datetime

import matplotlib

matplotlib.use("Qt5Agg")
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import re

"""
This file is to be used with AZMP survey data that does NOT have a switched sensor.
If a sensor has been switched, you will be prompted to run the "DO2_sensor_switch.py", please follow the instructions in the mentioned file to get proper excel output.

For this code you need the following files of trip data in the working directory:
    .bot, 
    meanO2.txt, 
    met.csv, 
    Instruments.csv

Please make sure you have the proper input when prompted so the correct files are being accessed, user input should match info in file names (i.e. no capital letters, proper year, ship code, etc.)

This file will create the following:
    joined.csv
    map.png
    plot_Sbeox0.png and plot_Sbeox1.png
    updated_data.csv
    DO2_calibration_output.xlsx

If you don't save BOTH graph images you will have issues with both the updated data files and the excel, the code will quit. Just run it again and make sure you save ALL graphs as images before pressing "Done" button!

If you are not satisfied with something in the excel output, make sure the excel file is CLOSED before running again.
You may need to delete the excel from the working directory if running the code does not immediately overwrite the excel.
In this case, the best thing to do is close excel, delete from working directory, run code, open excel, open your cd folder and select from there and NOT the main excel menu.

That should be all :)
"""

def generate_bot_df(path: str) -> pd.DataFrame:
    """
    This function generates a Pandas DataFrame from a CSV file located at the specified path.
    The generated DataFrame contains only a subset of columns.

    Args:


    Returns:
        pandas.DataFrame: The generated DataFrame containing the specified columns.

    Raises:
        FileNotFoundError: If the file at the specified path does not exist.
    """
    if not Path(path).exists():
        raise FileNotFoundError(
            f"Expected {path} to be in the current working directory, but it was not found"
        )

    orig_bot_df = pd.read_csv(path)
    bot_df = orig_bot_df.filter(
        items=[
            "Stickr",
            "Station-ID",
            "Date",
            "ShTrpStn",
            "latitude",
            "longitude",
            "PrDM",
            "Sbeox0ML/L",
            "Sbeox1ML/L",
        ]
    )

    bot_df.to_csv("botdf.csv")

    return orig_bot_df, bot_df


def generate_o2_df(path: str) -> pd.DataFrame:
    """
    This function generates a Pandas DataFrame from a text file located at the specified path.
    The header of the file is discarded, reads the rest of the file as CSV.
    The generated DataFrame contains only a subset of columns.

    Args:
        path (str): The file path to the text file containing the source data.

    Returns:
        pandas.DataFrame: The generated DataFrame containing the specified columns.

    Raises:
        FileNotFoundError: If the file at the specified path does not exist.
    """
    mean_o2_file_path = Path(path)
    if not mean_o2_file_path.exists():
        raise FileNotFoundError(
            f"Expected {path} to be in the current working directory, but it was not found"
        )

    mean_o2_file = mean_o2_file_path.read_text()
    _header, o2_csv = mean_o2_file.split("\n\n")
    orig_o2_df = pd.read_csv(StringIO(o2_csv))
    o2_df = orig_o2_df.filter(
        items=[
            "Sample",
            "[O2](ml/l)",
            "std_[O2](ml/l)",
            "Rep_1(ml/l)",
            "Rep_2(ml/l)",
            "Rep_3(ml/l)",
            "Rep_4(ml/l)",
            "Rep_5(ml/l)",
            "Rep_6(ml/l)",
        ]
    )

    return orig_o2_df, o2_df


def generate_met_df(path: str) -> pd.DataFrame:
    """
    This function generates a Pandas DataFrame containing weather data from a CSV file located at the specified path.
    Converts and renames the "Wind Speed" column from knots into metres per second.

    Args:
        path (str): A file path to the CSV file containing the weather data.

    Returns:
        pandas.DataFrame:
        A DataFrame containing the weather data filtered to include only the columns 'Station',
        'Wind Dir [deg/10]', and the converted 'Wind Speed [m/s]'.

    Raises:
        FileNotFoundError: If the file at the specified path does not exist.
    """
    if not Path(path).exists():
        raise FileNotFoundError(
            f"Expected {path} to be in the current working directory, but it was not found"
        )

    met_df = pd.read_csv(path)
    met_df = met_df.filter(items=["Station", "Wind Dir [deg/10]", "Wind Speed [kts]"])

    knot_to_mps_conversion = 1.944
    met_df["Wind Speed [m/s]"] = met_df["Wind Speed [kts]"].apply(
        lambda x: round(x / knot_to_mps_conversion, 4)
    )

    return met_df


def generate_instrument_df(path: str):
    """
    Generates a DataFrame from a CSV file containing instrument data.

    Args:
        path (str): The file path to the CSV file.

    Returns:
        pd.DataFrame: A DataFrame containing the instrument data.

    Raises:
        FileNotFoundError: If the specified file path does not exist.
    """
    if not Path(path).exists():
        raise FileNotFoundError(
            f"Expected {path} to be in the current working directory, but it was not found"
        )

    instr_df = pd.read_csv(path)

    return instr_df


def seperate_based_on_winkler_diff(
    df: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Separates the data into successful and failed titrations based on Winkler differences.

    Args:
        df (pd.DataFrame): DataFrame containing the titration data.

    Returns:
        Tuple[pd.DataFrame, pd.DataFrame]: A tuple of two DataFrames. The first DataFrame
            contains the data for successful titrations, and the second DataFrame contains
            the data for failed titrations.

    Raises:
        None

    """
    
    df["WINKLER-DIFF"] = df.apply(
        lambda x: abs(round(x["Rep_1(ml/l)"] - x["Rep_2(ml/l)"], 4)), axis=1
    )
    df["WINKLER"] = df.apply(
        lambda x: round((x["Rep_1(ml/l)"] + x["Rep_2(ml/l)"]) / 2, 4), axis=1
    )

    wanted_columns = ["Stickr", "Station-ID", "Sbeox0ML/L", "WINKLER", "WINKLER-DIFF"]

    successful_titrations_df = df.loc[
        (df["WINKLER"] <= 9) & (df["WINKLER"] >= 2) & (df["WINKLER-DIFF"] <= 0.2)
    ]

    failed_titrations_df = df.loc[
        (df["WINKLER"] > 9) | (df["WINKLER"] < 2) | (df["WINKLER-DIFF"] > 0.2),
        wanted_columns,
    ]


    return successful_titrations_df, failed_titrations_df


def join_bot_o2_met_dfs(
    bot_df: pd.DataFrame,
    o2_df: pd.DataFrame,
    met_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    This function joins three pandas DataFrames: bot_df, o2_df, and met_df then returns a merged DataFrame.

    Parameters:
    bot_df: pandas.DataFrame
        DataFrame containing the data of bot samples.
    o2_df: pandas.DataFrame
        DataFrame containing the O2 data.
    met_df: pandas.DataFrame
        DataFrame containing the data of meteorological samples.

    Returns:
    pandas.DataFrame
        A merged DataFrame of all three input DataFrames, joined on common columns:
        bot_df and met_df are joined on "ShTrpStn" and "Station".
        o2_df and join_dfs are joined on "Stickr" and "Sample".

    Raises:
        None.
    """

    join_dfs = bot_df.merge(met_df, how="inner", left_on="ShTrpStn", right_on="Station")
    join_dfs = join_dfs.merge(o2_df, how="inner", left_on="Stickr", right_on="Sample")
    join_dfs.drop(["Station", "Sample"], axis=1, inplace=True)

    return join_dfs


def generate_map(df: pd.DataFrame, bot_df: pd.DataFrame, title: str):
    """
    Generates a map plot showing the locations of oxygen samples and occupied stations.

    Args:
        df (pd.DataFrame): DataFrame containing the oxygen sample data.
        bot_df (pd.DataFrame): DataFrame containing the occupied station data.
        title (str): Title of the map plot.

    Returns:
        None

    Raises:
        None

    """
    plt.figure()

    m = Basemap(
        projection="merc",
        llcrnrlat=bot_df["latitude"].min() - 1.5,
        urcrnrlat=bot_df["latitude"].max() + 1.5,
        llcrnrlon=bot_df["longitude"].min() - 1.5,
        urcrnrlon=bot_df["longitude"].max() + 1.5,
        lat_ts=20,
        resolution="h",
    )

    x, y = m(df["longitude"].values, df["latitude"].values)

    all_x, all_y = m(bot_df["longitude"].values, bot_df["latitude"].values)

    m.drawcoastlines()
    m.scatter(all_x, all_y, s=7, color="blue", alpha=0.5, label="Occupied Stations")
    m.scatter(x, y, s=7, color="red", alpha=0.5, label="Oxygen Samples")
    plt.legend()
    plt.title("Filtered Results: {}".format(title))

    plt.savefig(f"{title}_map.png", dpi=100, bbox_inches="tight")


def new_and_old_SOC_calculations(instr_df: pd.DataFrame, joined_df: pd.DataFrame):
    """
    Calculate the new and old dissolved oxygen Standard Output Calibration (SOC) values from an instrument and a dataset.

    Args:
        instr_df (pd.DataFrame): A DataFrame containing the primary and secondary old dissolved oxygen SOC values.
        joined_df (pd.DataFrame): A DataFrame containing dissolved oxygen measurements and reference values.

    Returns:
        Tuple[List[float], float]: A tuple containing the new dissolved oxygen SOC values (as a list with primary and secondary
        values) and the averaged Winkler oxygen calculation.
    """

    df = joined_df.copy()
    winklerOx_calc = []
    newSOC = []

    df["WinkOxCalc0"] = df.apply(
        lambda x: round((x["WINKLER"] / x["Sbeox0ML/L"]), 4), axis=1
    )
    df["WinkOxCalc1"] = df.apply(
        lambda x: round((x["WINKLER"] / x["Sbeox1ML/L"]), 4), axis=1
    )

    oldSOC = [
        instr_df["Primary Oxygen Old SOC"].iloc[1],
        instr_df["Secondary Oxygen Old SOC"].iloc[1],
    ]
    winklerOx_calc.extend(df["WinkOxCalc0"].values.tolist())
    winklerOx_calc.extend(df["WinkOxCalc1"].values.tolist())
    winklerOx_avg = round(sum(winklerOx_calc) / len(winklerOx_calc), 4)

    newSOC.append(round(float((winklerOx_avg) * (oldSOC[0])), 4))
    newSOC.append(round(float((winklerOx_avg) * (oldSOC[1])), 4))

    return newSOC, oldSOC


def join_instr(joined_df:pd.DataFrame, instr_df: pd.DataFrame) -> pd.DataFrame:
    """
    Joins the data from the joined and instrument DataFrames based on a common key.

    Args:
        joined_df (pd.DataFrame): DataFrame containing the data to be joined.
        instr_df (pd.DataFrame): DataFrame containing the instrument data.

    Returns:
        pd.DataFrame: DataFrame resulting from the inner join of the two input DataFrames.

    Raises:
        None.

    """
    join_dfs = joined_df.merge(
        instr_df, how="inner", left_on="ShTrpStn", right_on="Filename"
    )
    join_dfs.drop(["Filename"], axis=1, inplace=True)

    return join_dfs


def filter_graph_data(joined_df: pd.DataFrame) -> pd.DataFrame:
    """
    Filters the joined dataframe + instruments to pass to the Plotter class.
    """
    graph_df = joined_df.filter(items=["Rep_1(ml/l)", "Sbeox0ML/L", "Sbeox1ML/L", "Primary Oxygen Serial Number", "Secondary Oxygen Serial Number"])

    return graph_df


class Plotter(QtWidgets.QWidget):
    def __init__(self, df):
        super().__init__()

        self.df = df.copy()
        self.original_df = df.copy()

        self.setWindowTitle("Plotter")
        self.setGeometry(100, 100, 800, 600)

        frame = QtWidgets.QFrame(self)
        frame_layout = QtWidgets.QHBoxLayout(frame)

        var_names = [
            c
            for c in df.columns
            if c
            not in [
                "index",
                "Rep_1(ml/l)",
                "Primary Oxygen Serial Number",
                "Secondary Oxygen Serial Number",
            ]
        ]
        self.var_selected = QtWidgets.QComboBox(self)
        self.var_selected.addItems(var_names)
        self.var_selected.currentIndexChanged.connect(self.update_plot)
        frame_layout.addWidget(self.var_selected)

        done_button = QtWidgets.QPushButton("Done", self)
        done_button.clicked.connect(self.close)
        frame_layout.addWidget(done_button)

        save_image_button = QtWidgets.QPushButton("Save Plot as Image", self)
        save_image_button.clicked.connect(self.save_image)
        frame_layout.addWidget(save_image_button)

        reset_button = QtWidgets.QPushButton("Reset Graph", self)
        reset_button.clicked.connect(self.reset_plot)
        frame_layout.addWidget(reset_button)

        layout = QtWidgets.QVBoxLayout(self)
        layout.addWidget(frame)

        self.figure = plt.figure()
        self.canvas = FigureCanvas(self.figure)

        layout.addWidget(self.canvas)
        self.canvas.ax = self.canvas.figure.subplots()

        self.rs = RectangleSelector(
            self.canvas.ax, self.onselect, interactive=True, useblit=True, button=[1]
        )

        self.update_plot()

    def closeEvent(self, event):
        self.df.to_csv(f"updated_data_{ship_code}{ship_trip}.csv", index=False)
        app = QtWidgets.QApplication.instance()
        app.quit()

    def determine_outliers(self):
        var = self.var_selected.currentText()

        x_q1 = self.df[var].quantile(0.25)
        x_q3 = self.df[var].quantile(0.75)
        x_iqr = x_q3 - x_q1

        y_q1 = self.df["Rep_1(ml/l)"].quantile(0.25)
        y_q3 = self.df["Rep_1(ml/l)"].quantile(0.75)
        y_iqr = y_q3 - y_q1

        upper_bound_x = x_q3 + 1.5 * x_iqr
        lower_bound_x = x_q1 - 1.5 * x_iqr

        upper_bound_y = y_q3 + 1.5 * y_iqr
        lower_bound_y = y_q1 - 1.5 * y_iqr

        outliers = self.df.loc[
            (self.df[var] < lower_bound_x)
            | (self.df[var] > upper_bound_x)
            | (self.df["Rep_1(ml/l)"] < lower_bound_y)
            | (self.df["Rep_1(ml/l)"] > upper_bound_y)
        ]

        return outliers

    def update_plot(self):
        var = self.var_selected.currentText()
        canvas = self.canvas
        ax = canvas.ax
        ax.clear()
        ax.scatter(
            self.df[var], self.df["Rep_1(ml/l)"], color="blue", label="Data Points"
        )

        outliers = self.determine_outliers()
        ax.scatter(
            outliers[var],
            outliers["Rep_1(ml/l)"],
            color="grey",
            label="Suggested Outliers",
        )

        idx = np.isfinite(self.df[var]) & np.isfinite(self.df["Rep_1(ml/l)"])
        z = np.polyfit(self.df[var][idx], self.df["Rep_1(ml/l)"][idx], 1)
        slope = z[0]
        intercept = z[1]
        p = np.poly1d(z)

        label = var.replace("ML/L", "")
        self.df[f"Slope_{label}"] = z[0]
        self.df[f"yint_{label}"] = z[1]

        self.df[f"Trendline_{label}"] = p(self.df[var])
        self.df = self.df.sort_values(by=[f"Trendline_{label}"])
        ax = self.df.plot(
            x=var,
            y=f"Trendline_{label}",
            color="red",
            ax=ax,
            label=f"Slope = {slope:.4f},  Y-Intercept: {intercept:.4f}",
        )
        
        if var == "Sbeox0ML/L":
            instrument_num_col = "Primary Oxygen Serial Number"

        if var == "Sbeox1ML/L":
            instrument_num_col = "Secondary Oxygen Serial Number"

        if instrument_num_col is not None:
            instrument_number = self.df.loc[1, instrument_num_col]
        else:
            instrument_number = "Unknown"

        self.canvas.ax.set_title(f"Instrument {instrument_number}")

        ax.legend()

        ax.set_xlabel(var)
        ax.set_ylabel("Rep_1(ml/l)")
        canvas.draw()

    def reset_plot(self):
        self.df = self.original_df
        self.update_plot()

    def save_image(self):
        var = self.var_selected.currentText()
        sample = var.replace("ML/L", "")
        self.figure.savefig(f"plot_{sample}_{graphic_titles}.png")

    def onselect(self, eclick, erelease):
        if eclick.ydata > erelease.ydata:
            eclick.ydata, erelease.ydata = erelease.ydata, eclick.ydata
            if eclick.xdata > erelease.xdata:
                eclick.xdata, erelease.xdata = erelease.xdata, eclick.xdata

        x1, y1 = eclick.xdata, eclick.ydata
        x2, y2 = erelease.xdata, erelease.ydata

        var = self.var_selected.currentText()
        selected_data = self.df[
            (self.df[var] >= x1)
            & (self.df[var] <= x2)
            & (self.df["Rep_1(ml/l)"] >= y1)
            & (self.df["Rep_1(ml/l)"] <= y2)
        ]

        self.df.loc[selected_data.index, var] = np.nan

        self.update_plot()


def generate_graph_df(path: str) -> pd.DataFrame:
    """
    Reads the updated graph CSV file and returns a DataFrame containing graph data.

    Args:
        path (str): The path to the CSV file.

    Returns:
        pd.DataFrame: The DataFrame containing the graph data.

    Raises:
        FileNotFoundError: If the specified file path does not exist.

    """
    if not Path(path).exists():
        raise FileNotFoundError(
            f"Expected {path} to be in the current working directory, but it was not found. Did you save both graphs?"
        )

    graph_df = pd.read_csv(path)

    return graph_df


def pressure_plot(graph_df: pd.DataFrame, joined_df):
    """
    TODO: write something here
    """
    
    y = joined_df["PrDM"]

    for i in range(2):
        plt.clf()
        x_orig = joined_df[f"Sbeox{i}ML/L"]
        x_cal = graph_df[f"Trendline_Sbeox{i}"]

        plt.scatter(x_orig, y, color="blue", label="Original")
        plt.scatter(x_cal, y, color="red", label ="Calibrated")
        plt.title(f"Original vs Calibrated Sbeox{i} values")
        plt.xlabel(f"Sbeox{i}")
        plt.ylabel("Pressure")
        plt.legend()
        plt.savefig(f"PrDM_plot{i}_{graphic_titles}.png")


def regression_data(
    instr_df: pd.DataFrame,
    graph_df: pd.DataFrame,
    newSOC: list,
    oldSOC: list,
):
    """
    This function takes in instrument data, graph data, newSOC, and oldSOC as inputs. It performs regression analysis
    using the graph data and extracts relevant information from the instrument data. The results are returned as a
    pandas DataFrame.

    Args:
        instr_df (pd.DataFrame): The instrument data DataFrame.
        graph_df (pd.DataFrame): The graph data DataFrame.
        newSOC (list): A list of newSOC values.
        oldSOC (list): A list of oldSOC values.

    Returns:
        pd.DataFrame: A DataFrame summarizing the regression analysis results, including instrument information,
        slope, y-intercept, oldSOC, and newSOC.

    Raises:
        None 
    """
    summary = pd.DataFrame

    primList = np.unique(instr_df["Primary Oxygen Serial Number"].values.tolist())
    secList = np.unique(instr_df["Secondary Oxygen Serial Number"])
    instr_list = np.append(primList, secList)
    slope = [graph_df["Slope_Sbeox0"].iloc[0], graph_df["Slope_Sbeox1"].iloc[0]]
    yint = [graph_df["yint_Sbeox0"].iloc[0], graph_df["yint_Sbeox1"].iloc[0]]

    summary = pd.DataFrame(
        {
            "Instr": instr_list,
            "Slope": slope,
            "y-int": yint,
            "OldSOC": oldSOC,
            "NewSOC": newSOC,
        }
    )

    return summary


def stations_occupied(df: pd.DataFrame):
    """
    Extracts the occupied station IDs from a DataFrame.

    Args:
        df (pd.DataFrame): The DataFrame containing station IDs.

    Returns:
        pd.DataFrame: A DataFrame with the extracted station IDs.
    """
    station_id = df["Station-ID"].drop_duplicates()
    station_id = np.unique(df["Station-ID"].tolist())
    station_dict = {}
    for station in station_id:
        val = re.split("-", station)
        if val[0] not in station_dict:
            station_dict[val[0]] = [val[1]]
        else:
            station_dict[val[0]].append(val[1])

    station_list = []
    for k in station_dict.keys():
        stat = (k, ", ".join(station_dict.get(k)))
        station_list.append(stat)

    station_df = pd.DataFrame(station_list)

    return station_df


def instruments_used(df: pd.DataFrame):
    """
    Extracts the unique oxygen instrument serial numbers used in a DataFrame.

    Args:
        df (pd.DataFrame): The DataFrame containing oxygen instrument serial numbers.

    Returns:
        Tuple[bool, pd.DataFrame]: A tuple containing a switch indicating if there is a switch in instruments
        and a DataFrame with the unique instrument serial numbers.
    """
    primList = np.unique(df["Primary Oxygen Serial Number"].values.tolist())
    secList = np.unique(df["Secondary Oxygen Serial Number"].values.tolist())

    switch = len(primList) > 1 or len(secList) > 1
     
    instr = pd.DataFrame(
        {
            "Primary": primList,
            "Secondary": secList,
        }
    )
    
    return switch, instr


def calibrated_dO2_sheet(orig_bot_df: pd.DataFrame, graph_df: pd.DataFrame) -> pd.DataFrame:
    """
    Calibrates dissolved oxygen measurements in the original bottle file using the slope of the updated graph data.

    Args:
        orig_bot_df (pd.DataFrame): The original bottle file DataFrame.
        graph_df (pd.DataFrame): The updated graph data DataFrame.

    Returns:
        pd.DataFrame: The calibrated bottle file DataFrame.
    """
    calibrated_df = orig_bot_df.copy()
    slope0 = graph_df["Slope_Sbeox0"].iloc[0]
    slope1 = graph_df["Slope_Sbeox1"].iloc[0]
    yint0 = graph_df["yint_Sbeox0"].iloc[0]
    yint1 = graph_df["yint_Sbeox1"].iloc[0]

    calibrated_df["Sbeox0_Calibrated"] = calibrated_df.apply(
        lambda x: round((x["Sbeox0ML/L"] * slope0) + yint0, 4), axis=1
    )
    calibrated_df["Sbeox1_Calibrated"] = calibrated_df.apply(
        lambda x: round((x["Sbeox1ML/L"] * slope1)+yint1, 4), axis=1)

    return calibrated_df


def write_to_DO2_excel(
    bot_file,
    txt_file,
    fail_tit,
    instr: pd.DataFrame,
    title: str,
    instr_used: pd.DataFrame,
    station_id: pd.DataFrame,
    reg_data: pd.DataFrame,
    dO2_cal: pd. DataFrame,
):
    start_date = bot_file["Date"].min()
    end_date = bot_file["Date"].max()

    wb = Workbook()

    # Summary sheet with map
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Oxygen Calibration"
    ws["A2"] = f"AZMP Survey for {ship_code} {ship_trip}"
    ws["A3"] = f"{start_date} to {end_date}"

    ws["A5"] = "The following instruments were used:"
    ws["A11"] = "The following stations were occupied:"
    ws["A18"] = "The regression data is:"

    map_img = Image(f"{title}_map.png")
    map_img.anchor = "I2"
    ws.add_image(map_img)

    ws = wb.create_sheet("Plots", 1)
    wb.active = 1
    plot0_png = Image(f"plot_Sbeox0_{graphic_titles}.png")
    plot0_png.anchor = "A1"
    plot1_png = Image(f"plot_Sbeox1_{graphic_titles}.png")
    plot1_png.anchor = "A30"
    ws.add_image(plot0_png)
    ws.add_image(plot1_png)
    wb.save(f"DO2_calibration_output_{title}.xlsx")
    with pd.ExcelWriter(
        f"DO2_calibration_output_{title}.xlsx",
        mode="a",
        engine="openpyxl",
        if_sheet_exists="overlay",
    ) as writer:
        instr_used.to_excel(
            writer, sheet_name="Summary", startrow=5, index=False, header=True
        )
        station_id.to_excel(
            writer, sheet_name="Summary", startrow=11, index=False, header=False
        )
        reg_data.to_excel(
            writer, sheet_name="Summary", startrow=18, index=False, header=True
        )
        # master_bot.to_excel(writer, sheet_name='Master bottle file')
        # all_data.to_excel(writer, sheet_name='Data')
        bot_file.to_excel(writer, sheet_name="Original bot file", index=False)
        dO2_cal.to_excel(writer, sheet_name="Calibrated bot file", index=False)
        txt_file.to_excel(writer, sheet_name="Original O2 text file", index=False)
        fail_tit.to_excel(writer, sheet_name="Failed Titrations", index=False)
        instr.to_excel(writer, sheet_name="Instruments", index=False)


def write_to_bot_excel(bot_file, met_file, title: str):
    """ 
    Not in use, creates the second excel with meteorological and bot information.
    """
    with pd.ExcelWriter(f"master_bot_output_{title}.xlsx", mode="w") as writer:
        met_file.to_excel(writer, sheet_name="met file", index=False)
        bot_file.to_excel(writer, sheet_name="Original bot file", index=False)


if __name__ == "__main__":
    year = int(input("Year: "))
    ship_code = input("Ship code: ")
    ship_trip = input("Ship trip: ")
    graphic_titles = f"{ship_code}{ship_trip}_{year}"

    instr_df = generate_instrument_df(f"{ship_code}{ship_trip}_{year}_Instruments.csv")
    switch, instr_used = instruments_used(instr_df)

    orig_bot_df, bot_df = generate_bot_df(f"{ship_code}{ship_trip}.bot")
    orig_o2_df, o2_df = generate_o2_df(f"{ship_code}{ship_trip}_meanO2.txt")
    met_df = generate_met_df(f"{ship_code}{ship_trip}_{year}_met.csv")

    joined_df = join_bot_o2_met_dfs(bot_df, o2_df, met_df)

    successful_titration, failed_titrations = seperate_based_on_winkler_diff(
        joined_df
    )

    joined_file = f"{ship_code}{ship_trip}_{year}_joined.csv"
    print(f"Written joined bot, o2 and met DataFrame to {joined_file}")
    successful_titration.to_csv(joined_file, index=False)

    generate_map(joined_df, bot_df, graphic_titles)

    if switch == True:
        print("A sensor has been switched, please run DO2_sensor_switch.py instead")
        exit(0)

    newSOC, oldSOC = new_and_old_SOC_calculations(instr_df, successful_titration)
    joined_instr_df = join_instr(joined_df, instr_df)

    # Generate + editing graph happens here
    graph_df = filter_graph_data(joined_instr_df)
    app = QtWidgets.QApplication([])
    plotter = Plotter(graph_df)
    plotter.show()
    app.exec_()
    app.quit()

    updated_graph_df = generate_graph_df(f"updated_data_{ship_code}{ship_trip}.csv")
    #prDMplt = pressure_plot(updated_graph_df, joined_df)
    station_id = stations_occupied(joined_df)
    reg_data = regression_data(instr_df, updated_graph_df, newSOC, oldSOC)
    dO2_cal = calibrated_dO2_sheet(orig_bot_df, updated_graph_df)

    write_to_DO2_excel(
        orig_bot_df,
        orig_o2_df,
        failed_titrations,
        instr_df,
        graphic_titles,
        instr_used,
        station_id,
        reg_data,
        dO2_cal,
    )
    write_to_bot_excel(orig_bot_df, met_df, graphic_titles)
