from pathlib import Path
from io import StringIO
from typing import Tuple
from matplotlib.widgets import RectangleSelector
from PyQt5 import QtWidgets
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
from mpl_toolkits.basemap import Basemap
from sklearn.metrics import r2_score

import matplotlib

matplotlib.use("Qt5Agg")
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import re
import DF_O2

"""
IMPORTANT: This code only works with trip data that has a switched sensor!!! If no switch go to DF_O2.py

This code should be excecuted as follows,

Make sure you have all modules above downloaded!!!

Make sure these essential files are in the current working directory BEFORE running this code!

.bot, meanO2.txt, met.csv, Instruments.csv, AND ***DF_O2.py***

Once those are in the working directory, you can run the code, my preference is to manually run it with "python .\DO2_sensor_switch.py"

Please make sure you have the proper input when prompted so the correct files are being accessed, user input should match info in file names (i.e. no capital letters, proper year, ship code, etc.)

As code runs, it will create files such as:
    joined.csv
    map.png
    plot_Sbeox0_0.png and plot_Sbeox1_0.png
    plot_Sbeox0_1.png and plot_Sbeox1_1.png
    updated_data0.csv and updated_data1.csv
    DO2_calibration_output.xlsx
 
If you don't save EVERY graph image you will have issues with both the updated data files and the excel, the code will quit. Just run it again and make sure you save ALL graphs as images before pressing "Done" button!

If you are not satisfied with something in the excel output, make sure the excel file is CLOSED before running again.
You may need to delete the excel from the working directory if running the code does not immediately overwrite the excel.
In this case, the best thing to do is close excel, delete from working directory, run code, open excel, open your cd folder and select from there and NOT the main excel menu.

That should be all :)
"""


def split_data(joined_df: pd.DataFrame, instr_df: pd.DataFrame) -> pd.DataFrame:
    """
    Splits the joined DataFrame based on the oxygen serial numbers.

    Args:
        joined_df (pd.DataFrame): DataFrame containing the joined data.
        instr_df (pd.DataFrame): DataFrame containing instrument data.

    Returns:
        tuple: A tuple containing two DataFrames:
            - df1: DataFrame filtered based on primary and secondary oxygen serial numbers.
            - anti_join: DataFrame with rows that do not match the primary oxygen serial number.

    Raises:
        None.
    """
    join_dfs = joined_df.merge(
        instr_df, how="inner", left_on="ShTrpStn", right_on="Filename"
    )
    join_dfs.drop(["Filename"], axis=1, inplace=True)

    df1 = join_dfs[
        join_dfs["Primary Oxygen Serial Number"]
        == instr_df["Primary Oxygen Serial Number"].iloc[0]
    ]
    df1 = df1[
        df1["Secondary Oxygen Serial Number"]
        == instr_df["Secondary Oxygen Serial Number"].iloc[0]
    ]

    df3 = join_dfs.merge(
        df1, on="Primary Oxygen Serial Number", how="left", indicator=True
    )

    df = df3.loc[df3["_merge"] == "left_only", "Primary Oxygen Serial Number"]

    anti_join = join_dfs[join_dfs["Primary Oxygen Serial Number"].isin(df)]

    anti_join.index = np.arange(0, len(anti_join))

    df1.to_csv("df1_sensor_switch.csv")
    anti_join.to_csv("df2_sensor_switch.csv")

    return df1, anti_join


def SOCcalc_split(split_df1: pd.DataFrame, split_df2: pd.DataFrame):
    """
    Calculates new SOC values based on oxygen measurements and old SOC values.
    Also determines the r2 value of graph data.

    Args:
        split_df1 (pd.DataFrame): DataFrame containing the first set of split data.
        split_df2 (pd.DataFrame): DataFrame containing the second set of split data.

    Returns:
        tuple: A tuple containing two lists:
            - newSOC: List of new SOC values calculated for each oxygen measurement.
            - oldSOC: List of old SOC values corresponding to the oxygen measurements.
            - r2_vals: List of r2 values.

    Raises:
        None.
    """

    df1 = split_df1.copy()
    df2 = split_df2.copy()

    newSOC = []
    oldSOC = []
    r2_vals = []

    df1["WinkOxCalc0"] = df1.apply(
        lambda x: round((x["WINKLER"] / x["Sbeox0ML/L"]), 4), axis=1
    )
    df1["WinkOxCalc1"] = df1.apply(
        lambda x: round((x["WINKLER"] / x["Sbeox1ML/L"]), 4), axis=1
    )

    df2["WinkOxCalc0"] = df2.apply(
        lambda x: round((x["WINKLER"] / x["Sbeox0ML/L"]), 4), axis=1
    )
    df2["WinkOxCalc1"] = df2.apply(
        lambda x: round((x["WINKLER"] / x["Sbeox1ML/L"]), 4), axis=1
    )

    df_list = [df1, df2]

    for i in df_list:
        oldSOC_val = [
            i["Primary Oxygen Old SOC"].iloc[1],
            i["Secondary Oxygen Old SOC"].iloc[1],
        ]

        winklerOx_calc = []
        winklerOx_calc.extend(i["WinkOxCalc0"].values.tolist())
        winklerOx_calc.extend(i["WinkOxCalc1"].values.tolist())
        winklerOx_avg = round(sum(winklerOx_calc) / len(winklerOx_calc), 4)

        oldSOC.extend([oldSOC_val[0], oldSOC_val[1]])
        newSOC.append(round(float((winklerOx_avg) * (oldSOC_val[0])), 4))
        newSOC.append(round(float((winklerOx_avg) * (oldSOC_val[1])), 4))

        r2_prim = r2_score(i["Sbeox0ML/L"], i["WINKLER"])
        r2_sec = r2_score(i["Sbeox1ML/L"], i["WINKLER"])

        r2_vals.extend([r2_prim, r2_sec])

    return newSOC, oldSOC, r2_vals


def instruments_used(df: pd.DataFrame):
    """
     Extracts the unique primary and secondary oxygen serial numbers used in the DataFrame.

    Args:
        df (pd.DataFrame): DataFrame containing instrument data.

    Returns:
        tuple: A tuple containing:
            - instr (pd.DataFrame): DataFrame with two columns: "Primary" and "Secondary"
              listing the unique primary and secondary oxygen serial numbers.
            - primList (list): List of unique primary oxygen serial numbers.
            - secList (list): List of unique secondary oxygen serial numbers.

    Raises:
        None.
    """
    prim, ind1 = np.unique(
        df["Primary Oxygen Serial Number"].values.tolist(), return_index=True
    )
    sec, ind2 = np.unique(
        df["Secondary Oxygen Serial Number"].values.tolist(), return_index=True
    )

    primList = prim[np.argsort(ind1)]
    secList = sec[np.argsort(ind2)]

    instr = pd.DataFrame(
        {
            "Primary": primList,
            "Secondary": secList,
        }
    )

    return instr, primList, secList


class Plotter(QtWidgets.QWidget):
    def __init__(self, df):
        super().__init__()

        self.df = df.copy()
        self.original_df = df.copy()

        self.setWindowTitle("Plotter")
        self.setGeometry(100, 100, 800, 600)

        frame = QtWidgets.QFrame(self)
        frame_layout = QtWidgets.QHBoxLayout(frame)

        var_names = [
            c
            for c in df.columns
            if c
            not in [
                "index",
                "Rep_1(ml/l)",
                "Primary Oxygen Serial Number",
                "Secondary Oxygen Serial Number",
            ]
        ]
        self.var_selected = QtWidgets.QComboBox(self)
        self.var_selected.addItems(var_names)
        self.var_selected.currentIndexChanged.connect(self.update_plot)
        frame_layout.addWidget(self.var_selected)

        done_button = QtWidgets.QPushButton("Done", self)
        done_button.clicked.connect(self.close)
        frame_layout.addWidget(done_button)

        save_image_button = QtWidgets.QPushButton("Save Plot as Image", self)
        save_image_button.clicked.connect(self.save_image)
        frame_layout.addWidget(save_image_button)

        reset_button = QtWidgets.QPushButton("Reset Graph", self)
        reset_button.clicked.connect(self.reset_plot)
        frame_layout.addWidget(reset_button)

        layout = QtWidgets.QVBoxLayout(self)
        layout.addWidget(frame)

        self.figure = plt.figure()
        self.canvas = FigureCanvas(self.figure)

        layout.addWidget(self.canvas)
        self.canvas.ax = self.canvas.figure.subplots()

        self.rs = RectangleSelector(
            self.canvas.ax, self.onselect, interactive=True, useblit=True, button=[1]
        )

        self.update_plot()

    def closeEvent(self, event):
        self.df.to_csv(f"updated_data_{ship_code}{ship_trip}{df_num}.csv", index=False)
        app = QtWidgets.QApplication.instance()
        app.quit()

    def determine_outliers(self):
        var = self.var_selected.currentText()

        x_q1 = self.df[var].quantile(0.25)
        x_q3 = self.df[var].quantile(0.75)
        x_iqr = x_q3 - x_q1

        y_q1 = self.df["Rep_1(ml/l)"].quantile(0.25)
        y_q3 = self.df["Rep_1(ml/l)"].quantile(0.75)
        y_iqr = y_q3 - y_q1

        upper_bound_x = x_q3 + 1.5 * x_iqr
        lower_bound_x = x_q1 - 1.5 * x_iqr

        upper_bound_y = y_q3 + 1.5 * y_iqr
        lower_bound_y = y_q1 - 1.5 * y_iqr

        outliers = self.df.loc[
            (self.df[var] < lower_bound_x)
            | (self.df[var] > upper_bound_x)
            | (self.df["Rep_1(ml/l)"] < lower_bound_y)
            | (self.df["Rep_1(ml/l)"] > upper_bound_y)
        ]

        return outliers

    def update_plot(self):
        var = self.var_selected.currentText()
        canvas = self.canvas
        ax = canvas.ax
        ax.clear()
        ax.scatter(
            self.df[var], self.df["Rep_1(ml/l)"], color="blue", label="Data Points"
        )

        outliers = self.determine_outliers()
        ax.scatter(
            outliers[var],
            outliers["Rep_1(ml/l)"],
            color="grey",
            label="Suggested Outliers",
        )

        idx = np.isfinite(self.df[var]) & np.isfinite(self.df["Rep_1(ml/l)"])
        z = np.polyfit(self.df[var][idx], self.df["Rep_1(ml/l)"][idx], 1)
        slope = z[0]
        intercept = z[1]
        p = np.poly1d(z)

        label = var.replace("ML/L", "")
        self.df[f"Slope_{label}"] = z[0]
        self.df[f"yint_{label}"] = z[1]

        self.df[f"Trendline_{label}"] = p(self.df[var])
        self.df = self.df.sort_values(by=[f"Trendline_{label}"])
        ax = self.df.plot(
            x=var,
            y=f"Trendline_{label}",
            color="red",
            ax=ax,
            label=f"Slope = {slope:.4f},  Y-Intercept: {intercept:.4f}",
        )

        if var == "Sbeox0ML/L":
            instrument_num_col = "Primary Oxygen Serial Number"

        if var == "Sbeox1ML/L":
            instrument_num_col = "Secondary Oxygen Serial Number"

        if instrument_num_col is not None:
            instrument_number = self.df.loc[1, instrument_num_col]
        else:
            instrument_number = "Unknown"

        self.canvas.ax.set_title(f"Instrument {instrument_number}")

        ax.legend()

        ax.set_xlabel(var)
        ax.set_ylabel("Rep_1(ml/l)")
        canvas.draw()

    def reset_plot(self):
        self.df = self.original_df
        self.update_plot()

    def save_image(self):
        var = self.var_selected.currentText()
        sample = var.replace("ML/L", "")
        self.figure.savefig(f"plot_{sample}_{graphic_titles}_{df_num}.png")

    def onselect(self, eclick, erelease):
        if eclick.ydata > erelease.ydata:
            eclick.ydata, erelease.ydata = erelease.ydata, eclick.ydata
            if eclick.xdata > erelease.xdata:
                eclick.xdata, erelease.xdata = erelease.xdata, eclick.xdata

        x1, y1 = eclick.xdata, eclick.ydata
        x2, y2 = erelease.xdata, erelease.ydata

        var = self.var_selected.currentText()
        selected_data = self.df[
            (self.df[var] >= x1)
            & (self.df[var] <= x2)
            & (self.df["Rep_1(ml/l)"] >= y1)
            & (self.df["Rep_1(ml/l)"] <= y2)
        ]

        self.df.loc[selected_data.index, var] = np.nan

        self.update_plot()


def updated_graph_dfs(path1: str, path2: str):
    """
    Reads and returns two DataFrames from the specified CSV files.

    Args:
        path1 (str): Path to the first CSV file.
        path2 (str): Path to the second CSV file.

    Returns:
        tuple: A tuple containing two DataFrames:
            - graphA_df (pd.DataFrame): DataFrame read from the first CSV file.
            - graphB_df (pd.DataFrame): DataFrame read from the second CSV file.

    Raises:
        FileNotFoundError: If either path1 or path2 does not exist in the current working directory.
    """
    if not Path(path1).exists() or not Path(path2).exists:
        raise FileNotFoundError(
            f"{path1} and {path2} required to be in the current working directory. Did you save both graphs?"
        )

    graphA_df = pd.read_csv(path1)
    graphB_df = pd.read_csv(path2)

    return graphA_df, graphB_df


def regression_data(
    graphA_df: pd.DataFrame,
    graphB_df: pd.DataFrame,
    newSOC: list,
    oldSOC: list,
    r2_vals: list,
):
    """
    Generates a summary DataFrame with regression data, SOC values, and instrument information.

    Args:
        graphA_df (pd.DataFrame): DataFrame containing regression data from graph A.
        graphB_df (pd.DataFrame): DataFrame containing regression data from graph B.
        newSOC (list): List of new SOC values.
        oldSOC (list): List of old SOC values.
        r2_vals (list): List of r2 values.

    Returns:
        pd.DataFrame: Summary DataFrame with the following columns:
            - "Instr": Instrument information.
            - "Slope": Slope values from graph A and graph B.
            - "y-int": Y-intercept values from graph A and graph B.
            - "r2 value: r2 values from graph A and graph B.
            - "OldSOC": Old SOC values.
            - "NewSOC": New SOC values.

    Raises:
        None.
    """

    summary = pd.DataFrame
    instr_list = []

    for i in range(0, len(primList)):
        instr_list.append(primList[i])
        instr_list.append(secList[i])

    slope = [
        graphA_df["Slope_Sbeox0"].iloc[0],
        graphA_df["Slope_Sbeox1"].iloc[0],
        graphB_df["Slope_Sbeox0"].iloc[0],
        graphB_df["Slope_Sbeox1"].iloc[0],
    ]
    yint = [
        graphA_df["yint_Sbeox0"].iloc[0],
        graphA_df["yint_Sbeox1"].iloc[0],
        graphB_df["yint_Sbeox0"].iloc[0],
        graphB_df["yint_Sbeox1"].iloc[0],
    ]

    summary = pd.DataFrame(
        {
            "Instr": instr_list,
            "Slope": slope,
            "y-int": yint,
            "r2 value": r2_vals,
            "OldSOC": oldSOC,
            "NewSOC": newSOC,
        }
    )

    return summary


def find_split_for_calibration(instr_df: pd.DataFrame, orig_bot_df: pd.DataFrame):
    df1 = instr_df[
        instr_df["Primary Oxygen Serial Number"]
        == instr_df["Primary Oxygen Serial Number"].iloc[0]
    ]
    df1 = df1[
        df1["Secondary Oxygen Serial Number"]
        == instr_df["Secondary Oxygen Serial Number"].iloc[0]
    ]

    split_point = df1["Filename"].iloc[-1]
    stop_num = int(split_point[-3:])
    orig_bot_df["ShTrpStn num"] = (
        orig_bot_df["ShTrpStn"].astype(str).apply(lambda x: x[-3:]).astype(int)
    )

    bot_df1 = orig_bot_df[orig_bot_df["ShTrpStn num"] <= stop_num].copy()
    bot_df2 = orig_bot_df[orig_bot_df["ShTrpStn num"] > stop_num].copy()

    bot_df1.drop(["ShTrpStn num"], axis=1, inplace=True)
    bot_df2.drop(["ShTrpStn num"], axis=1, inplace=True)

    return bot_df1, bot_df2


def calibrated_dO2_sheet(
    orig_bot_df: pd.DataFrame,
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    graphA_df: pd.DataFrame,
    graphB_df: pd.DataFrame,
) -> pd.DataFrame:
    """ """
    slope0A = graphA_df["Slope_Sbeox0"].iloc[0]
    slope1A = graphA_df["Slope_Sbeox1"].iloc[0]
    slope0B = graphB_df["Slope_Sbeox0"].iloc[0]
    slope1B = graphB_df["Slope_Sbeox1"].iloc[0]

    yint0A = graphA_df["yint_Sbeox0"].iloc[0]
    yint1A = graphA_df["yint_Sbeox1"].iloc[0]
    yint0B = graphB_df["yint_Sbeox0"].iloc[0]
    yint1B = graphB_df["yint_Sbeox1"].iloc[0]

    df1["Sbeox0_Calibrated"] = df1.apply(
        lambda x: round((x["Sbeox0ML/L"] * slope0A) + yint0A, 4), axis=1
    )

    df1["Sbeox1_Calibrated"] = df1.apply(
        lambda x: round((x["Sbeox1ML/L"] * slope1A) + yint1A, 4), axis=1
    )

    df2["Sbeox0_Calibrated"] = df2.apply(
        lambda x: round((x["Sbeox0ML/L"] * slope0B) + yint0B, 4), axis=1
    )

    df2["Sbeox1_Calibrated"] = df2.apply(
        lambda x: round((x["Sbeox1ML/L"] * slope1B) + yint1B, 4), axis=1
    )

    calibrated_df = pd.concat([df1, df2], join="outer", ignore_index=True)

    return calibrated_df


def write_to_DO2_excel(
    bot_file,
    txt_file,
    fail_tit,
    instr: pd.DataFrame,
    title: str,
    instr_used: pd.DataFrame,
    station_id: pd.DataFrame,
    reg_data: pd.DataFrame,
    dO2_cal: pd.DataFrame,
):
    """
    Writes the calibration results and related data to an Excel file.

    Args:
        bot_file (pd.DataFrame): DataFrame containing the original bot file data.
        txt_file (pd.DataFrame): DataFrame containing the O2.txt file data.
        fail_tit (pd.DataFrame): DataFrame containing the failed titrations data.
        instr (pd.DataFrame): DataFrame containing instrument data.
        title (str): Title of the Excel file.
        instr_used (pd.DataFrame): DataFrame containing the list of instruments used.
        station_id (pd.DataFrame): DataFrame containing the list of occupied stations.
        reg_data (pd.DataFrame): DataFrame containing the regression data.
        dO2_cal (pd.DataFrame): DataFrame containing the calibrated bot file data.

    Returns:
        None. The calibration results are written to an Excel file.

    Raises:
        None.

    """

    start_date = bot_file["Date"].min()
    end_date = bot_file["Date"].max()

    wb = Workbook()

    # Summary sheet with map
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Oxygen Calibration"
    ws["A2"] = f"AZMP Survey for {ship_code} {ship_trip}"
    ws["A3"] = f"{start_date} to {end_date}"

    ws["A5"] = "The following instruments were used:"
    ws["A11"] = "The following stations were occupied:"
    ws["A18"] = "The regression data is:"

    map_img = Image(f"{title}_map.png")
    map_img.anchor = "I2"
    ws.add_image(map_img)

    ws = wb.create_sheet("Plots", 1)
    wb.active = 1
    plot0A_png = Image(f"plot_Sbeox0_{graphic_titles}_0.png")
    plot0A_png.anchor = "A1"
    plot1A_png = Image(f"plot_Sbeox1_{graphic_titles}_0.png")
    plot1A_png.anchor = "A30"
    ws.add_image(plot0A_png)
    ws.add_image(plot1A_png)
    plot0B_png = Image(f"plot_Sbeox0_{graphic_titles}_1.png")
    plot0B_png.anchor = "O1"
    plot1B_png = Image(f"plot_Sbeox1_{graphic_titles}_1.png")
    plot1B_png.anchor = "O30"
    ws.add_image(plot0B_png)
    ws.add_image(plot1B_png)

    wb.save(f"DO2_calibration_output_{title}.xlsx")
    with pd.ExcelWriter(
        f"DO2_calibration_output_{title}.xlsx",
        mode="a",
        engine="openpyxl",
        if_sheet_exists="overlay",
    ) as writer:
        instr_used.to_excel(
            writer, sheet_name="Summary", startrow=5, index=False, header=True
        )
        station_id.to_excel(
            writer, sheet_name="Summary", startrow=11, index=False, header=False
        )
        reg_data.to_excel(
            writer, sheet_name="Summary", startrow=18, index=False, header=True
        )

        bot_file.to_excel(writer, sheet_name="Original bot file", index=False)
        dO2_cal.to_excel(writer, sheet_name="Calibrated bot file", index=False)
        txt_file.to_excel(writer, sheet_name="O2.txt file", index=False)
        fail_tit.to_excel(writer, sheet_name="Failed Titrations", index=False)
        instr.to_excel(writer, sheet_name="Instruments", index=False)


def write_to_bot_excel(bot_file, met_file, title: str):
    """
    Creates bot excel, in progress
    """
    with pd.ExcelWriter(f"master_bot_output_{title}.xlsx", mode="w") as writer:
        met_file.to_excel(writer, sheet_name="met file", index=False)
        bot_file.to_excel(writer, sheet_name="Original bot file", index=False)


if __name__ == "__main__":
    year = int(input("Year: "))
    ship_code = input("Ship code: ")
    ship_trip = input("Ship trip: ")
    graphic_titles = f"{ship_code}{ship_trip}_{year}"

    instr_df = DF_O2.generate_instrument_df(
        f"{ship_code}{ship_trip}_{year}_Instruments.csv"
    )
    instr_used, primList, secList = instruments_used(instr_df)

    orig_bot_df, bot_df = DF_O2.generate_bot_df(f"{ship_code}{ship_trip}.bot")
    orig_o2_df, o2_df = DF_O2.generate_o2_df(f"{ship_code}{ship_trip}_meanO2.txt")
    met_df = DF_O2.generate_met_df(f"{ship_code}{ship_trip}_{year}_met.csv")

    joined_df = DF_O2.join_bot_o2_met_dfs(bot_df, o2_df, met_df)

    successful_titration, failed_titrations = DF_O2.seperate_based_on_winkler_diff(
        joined_df
    )

    joined_file = f"{ship_code}{ship_trip}_{year}_joined.csv"
    successful_titration.to_csv(joined_file, index=False)
    print(f"Written joined bot, o2 and met DataFrame to {joined_file}")

    station_id = DF_O2.stations_occupied(joined_df)
    DF_O2.generate_map(joined_df, bot_df, graphic_titles)

    df1, df2 = split_data(successful_titration, instr_df)
    newSOC, oldSOC, r2_vals = SOCcalc_split(df1, df2)

    df_num = 0
    graphA_df = DF_O2.filter_graph_data(df1)
    app = QtWidgets.QApplication([])
    plotter = Plotter(graphA_df)
    plotter.show()
    app.exec_()

    df_num += 1
    graphB_df = DF_O2.filter_graph_data(df2)
    plotter = Plotter(graphB_df)
    plotter.show()
    app.exec_()
    app.quit()

    graphA_df, graphB_df = updated_graph_dfs(
        f"updated_data_{ship_code}{ship_trip}0.csv",
        f"updated_data_{ship_code}{ship_trip}1.csv",
    )
    # prDMplt = pressure_plot(updated_graph_df, joined_df) PRESSURE PLOT, not currently being used, does not exist in this code yet

    split_df1, split_df2 = find_split_for_calibration(instr_df, orig_bot_df)
    reg_data = regression_data(graphA_df, graphB_df, newSOC, oldSOC, r2_vals)
    bot_df1, bot_df2 = find_split_for_calibration(instr_df, orig_bot_df)
    dO2_cal = calibrated_dO2_sheet(
        orig_bot_df, split_df1, split_df2, graphA_df, graphB_df
    )

    write_to_DO2_excel(
        orig_bot_df,
        orig_o2_df,
        failed_titrations,
        instr_df,
        graphic_titles,
        instr_used,
        station_id,
        reg_data,
        dO2_cal,
    )
    # write_to_bot_excel(orig_bot_df, met_df, graphic_titles). not being used, creates the second excel file
